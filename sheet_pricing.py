"""Read quote/steel sheet dimensions and CSV prices for a job.

Prices come only from CSV files (never guessed):
  - Shop: Purchased Components Prices.csv
  - Per-job: Purchased Components Quote.csv (written by Module6121)
  - Per-job: Pullcore Prices.csv

Steel/quote Excel workbooks supply stock/finished sizing for plates.
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

from . import config

# Role -> keywords to match Component column in price CSVs
ROLE_CSV_KEYWORDS: dict[str, list[str]] = {
    "top_clamp_plate": ["top clamp", "top plate", "top clamping"],
    "a_plate": ["a plate", "a-plate"],
    "b_plate": ["b plate", "b-plate"],
    "stripper_plate": ["stripper"],
    "sc_retainer_plate": ["sc retainer", "retainer plate"],
    "sc_backup_plate": ["sc backup", "backup plate"],
    "support_plate": ["support plate"],
    "bottom_clamp_plate": ["bottom clamp", "bottom plate", "bot clamp"],
    "rail": ["rail"],
    "pin_plate": ["pin plate"],
    "ejector_plate": ["ejector plate", "ej-ret", "ej ret"],
    "bottom_ejector_plate": ["bottom ejector", "ej-backup", "ej backup", "ejector backup"],
    "latch_lock": ["safety strap", "latch lock", "latch-lock", "lss"],
    "leader_pin": ["leader pin", "ldr-pin", "ldr pin"],
    "leader_pin_bushing": ["bushing", "leader bushing", "guide bushing", "lbb"],
    "guided_ejector_bushing": ["ejector bushing"],
    "return_pin": ["return pin"],
    "ejector_pin": ["ejector pin"],
    "support_pillar": ["support pillar", "pillar"],
    "pullcore": ["pullcore", "pull core"],
}

PLATE_SHEET_NAMES: dict[str, list[str]] = {
    "top_clamp_plate": ["top clamp plate", "top clamping plate"],
    "a_plate": ["a plate"],
    "b_plate": ["b plate"],
    "stripper_plate": ["stripper plate"],
    "sc_retainer_plate": ["sc retainer plate", "sc retainer"],
    "sc_backup_plate": ["sc backup plate", "sc backup"],
    "support_plate": ["support plate"],
    "bottom_clamp_plate": ["bottom clamp plate", "bottom clamping plate"],
    "rail": ["rail"],
    "pin_plate": ["pin plate"],
    "ejector_plate": ["ejector plate"],
    "bottom_ejector_plate": ["bottom ejector plate", "ejector backup plate"],
}


def _safe_float(v, default=0.0) -> float:
    try:
        s = str(v).strip().replace("$", "").replace(",", "")
        return float(s) if s else default
    except (TypeError, ValueError):
        return default


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").lower().strip())


def _read_csv_rows(path: Path) -> list[dict]:
    if not path.exists():
        return []
    rows = []
    try:
        with path.open("r", newline="", encoding="utf-8-sig", errors="replace") as f:
            for line in f:
                if line.lstrip().startswith("#"):
                    continue
                rows.append(line)
        if not rows:
            return []
        reader = csv.DictReader(rows)
        return list(reader)
    except Exception:
        return []


def load_shop_prices() -> list[dict]:
    return _read_csv_rows(config.PURCHASED_PRICES_CSV)


def load_job_purchased_quote(job_dir: Path) -> list[dict]:
    for name in (
        "Purchased Components Quote.csv",
        "documents/Purchased Components Quote.csv",
    ):
        p = job_dir / name
        rows = _read_csv_rows(p)
        if rows:
            return rows
    return []


def load_job_pullcore(job_dir: Path) -> list[dict]:
    for name in ("Pullcore Prices.csv", "documents/Pullcore Prices.csv"):
        p = job_dir / name
        rows = _read_csv_rows(p)
        if rows:
            return rows
    return []


# BMS QuoteWorksheet #2 4140 block (Module6121 FillQuoteWorkbookFromBoundingBox).
# Col A = name, C = qty, D = thickness, E = width, F = length.
# Template formulas often put hours in G and price in H.
# Rows 22/23/31-34 are the canonical pot-block plates; 24-30 are spare/extra.
_BMS_QUOTE_STEEL_ROWS = {
    22: ("TCP", "tcp"),
    23: ("BCP", "bcp"),
    24: ("", "steel_plate"),
    25: ("", "steel_plate"),
    26: ("", "steel_plate"),
    27: ("", "steel_plate"),
    28: ("", "steel_plate"),
    29: ("", "steel_plate"),
    30: ("", "steel_plate"),
    31: ("ID Holder", "id_holder"),
    32: ("OD Holder", "od_holder"),
    33: ("ID Pot", "id_pot"),
    34: ("OD Pot", "od_pot"),
}

_BMS_NAME_TO_ROLE = {
    "tcp": "tcp",
    "bcp": "bcp",
    "id holder": "id_holder",
    "od holder": "od_holder",
    "id pot": "id_pot",
    "od pot": "od_pot",
    "id pot block": "id_pot",
    "od pot block": "od_pot",
    "id holder block": "id_holder",
    "od holder block": "od_holder",
}

_BMS_ROLE_LABELS = {
    "tcp": "TCP",
    "bcp": "BCP",
    "id_holder": "ID Holder",
    "od_holder": "OD Holder",
    "id_pot": "ID Pot",
    "od_pot": "OD Pot",
    "steel_plate": "Steel Plate",
}


def _cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _canonical_steel_name(name: str, role: str, default_name: str = "") -> str:
    """Never return a blank steel description — UI shows '--' otherwise."""
    n = _cell_str(name)
    if n and n not in {"--", "-", "None", "none"}:
        return n
    if default_name:
        return default_name
    return _BMS_ROLE_LABELS.get(role, "") or role or "Steel Plate"


def _iter_workbook_rows(wb_path: Path, sheet_names: tuple[str, ...] | None = None):
    """Yield (sheet_name, 1-based_row_index, cell_values_list)."""
    suffix = wb_path.suffix.lower()
    if suffix == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError:
            return
        try:
            book = xlrd.open_workbook(str(wb_path))
        except Exception:
            return
        for name in book.sheet_names():
            if sheet_names and name not in sheet_names:
                low = name.lower()
                if not any(s.lower() in low for s in sheet_names):
                    continue
            sh = book.sheet_by_name(name)
            for r in range(sh.nrows):
                yield name, r + 1, list(sh.row_values(r))
        return

    try:
        import openpyxl  # type: ignore
    except ImportError:
        return
    try:
        # data_only=True needs Excel to have calculated+saved formulas.
        # Fall back to formula workbook if cached values are missing.
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    except Exception:
        return
    try:
        for name in wb.sheetnames:
            if sheet_names and name not in sheet_names:
                low = name.lower()
                if not any(s.lower() in low for s in sheet_names):
                    continue
            ws = wb[name]
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=280, values_only=True), start=1):
                yield name, r_idx, list(row) if row else []
    finally:
        wb.close()


def _score_quote_workbook(path: Path) -> int:
    nm = path.name.upper()
    if "PURCHASED" in nm:
        return -1000
    score = 0
    if "QUOTE" in nm and "STEEL" in nm:
        score += 80
    if "GRIND" in nm:
        score += 30
    if "QUOTE" in nm:
        score += 20
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        score += 5
    # Prefer shorter names (fewer date/initials suffixes).
    score -= max(0, len(path.name) - 40) // 5
    return score


def _score_steel_workbook(path: Path) -> int:
    nm = path.name.upper()
    if "PURCHASED" in nm:
        return -1000
    score = 0
    if "STEEL" in nm and "SHEET" in nm:
        score += 80
    if "J000" in nm:
        score += 40
    if "MACHINING" in nm:
        score += 20
    # Quote_Steel_Grinding also matches *steel* — de-prioritize vs true J000.
    if "QUOTE" in nm and "GRIND" in nm:
        score -= 30
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        score += 5
    score -= max(0, len(path.name) - 40) // 5
    return score


def _iter_candidate_workbooks(job_dir: Path):
    seen: set[str] = set()
    for sub in ("", "documents"):
        base = job_dir / sub if sub else job_dir
        if not base.exists():
            continue
        for pat in ("*.xlsx", "*.xlsm", "*.xls"):
            for hit in sorted(base.glob(pat)):
                key = hit.name.lower()
                if key.startswith("~$"):
                    continue
                if key in seen:
                    continue
                seen.add(key)
                yield hit


def _find_workbook(job_dir: Path) -> Path | None:
    """Best quote workbook (Quote_Steel_Grinding), else best steel sheet."""
    quote = _find_quote_workbook(job_dir)
    if quote:
        return quote
    return _find_steel_workbook(job_dir)


def _find_quote_workbook(job_dir: Path) -> Path | None:
    best: Path | None = None
    best_score = 0
    for hit in _iter_candidate_workbooks(job_dir):
        sc = _score_quote_workbook(hit)
        if sc > best_score:
            best_score = sc
            best = hit
    return best if best_score > 0 else None


def _find_steel_workbook(job_dir: Path) -> Path | None:
    best: Path | None = None
    best_score = 0
    for hit in _iter_candidate_workbooks(job_dir):
        sc = _score_steel_workbook(hit)
        if sc > best_score:
            best_score = sc
            best = hit
    return best if best_score > 0 else None


def load_steel_plate_lines(job_dir: Path) -> list[dict]:
    """Steel plates from J000 (names) merged with QuoteWorksheet hours/price.

    Module6121 writes plate *names* reliably on the J000 Steel Order sheet.
    The QuoteWorksheet #2 block has stock dims + Excel hours/price formulas
    (only populated after Excel calculates). Prefer J000 for the row list so
    the UI never shows blank Description='--'.
    """
    quote_wb = _find_quote_workbook(job_dir)
    steel_wb = _find_steel_workbook(job_dir)

    j000_lines = _load_steel_from_j000(steel_wb) if steel_wb else []
    # Quote workbook sometimes *is* the only file; also try J000 sheets there.
    if not j000_lines and quote_wb and quote_wb != steel_wb:
        j000_lines = _load_steel_from_j000(quote_wb)

    quote_lines = _load_steel_from_quote_block(quote_wb) if quote_wb else []

    if j000_lines:
        return _merge_steel_hours_prices(j000_lines, quote_lines)

    if quote_lines:
        # Ensure every line has a visible name.
        for line in quote_lines:
            line["component"] = _canonical_steel_name(
                line.get("component") or "",
                line.get("role") or "",
            )
        return quote_lines

    return _load_steel_from_bom_match(job_dir)


def _merge_steel_hours_prices(base_lines: list[dict], quote_lines: list[dict]) -> list[dict]:
    """Copy hours/price from quote block onto J000 rows matched by role (then dims)."""
    if not quote_lines:
        return base_lines

    by_role: dict[str, list[dict]] = {}
    for q in quote_lines:
        role = q.get("role") or ""
        by_role.setdefault(role, []).append(q)

    used: set[int] = set()
    out: list[dict] = []
    for line in base_lines:
        merged = dict(line)
        merged["component"] = _canonical_steel_name(
            merged.get("component") or "",
            merged.get("role") or "",
        )
        role = merged.get("role") or ""
        candidates = by_role.get(role) or []
        best = None
        best_idx = -1
        best_score = -1.0
        for i, q in enumerate(candidates):
            if id(q) in used:
                continue
            score = 0.0
            for key in ("thickness", "width", "length"):
                a = float(merged.get(key) or 0)
                b = float(q.get(key) or 0)
                if a > 0 and b > 0 and abs(a - b) < 0.35:
                    score += 1.0
            if float(q.get("price") or 0) > 0:
                score += 0.25
            if float(q.get("hours") or 0) > 0:
                score += 0.25
            if score > best_score:
                best_score = score
                best = q
                best_idx = i
        if best is not None and best_score >= 1.0:
            used.add(id(best))
            if best.get("hours"):
                merged["hours"] = best["hours"]
            if float(best.get("price") or 0) > 0:
                merged["price"] = best["price"]
                merged["price_source"] = best.get("price_source") or merged.get("price_source")
            # Keep J000 finished dims (source of truth for T/W/L). Quote stock
            # sizes can swap axes vs the steel sheet — do not overwrite.
        out.append(merged)
    return out


def _load_steel_from_quote_block(wb_path: Path | None) -> list[dict]:
    if not wb_path:
        return []
    lines: list[dict] = []
    for sheet_name, row_idx, cells in _iter_workbook_rows(
        wb_path, ("QuoteWorksheet", "Quote")
    ):
        # Prefer the real QuoteWorksheet; skip fuzzy "Quote Summary" etc. when
        # the exact sheet exists in this workbook iteration order.
        if sheet_name not in ("QuoteWorksheet", "Quote") and "quote" not in sheet_name.lower():
            continue
        if row_idx not in _BMS_QUOTE_STEEL_ROWS:
            continue
        default_name, role = _BMS_QUOTE_STEEL_ROWS[row_idx]
        while len(cells) < 10:
            cells.append(None)

        # Col A preferred; some templates put the label in Col B.
        raw_name = _cell_str(cells[0]) or _cell_str(cells[1])
        name = _canonical_steel_name(raw_name, role, default_name)
        qty = _safe_float(cells[2], 0.0)
        thickness = _parse_fraction(cells[3])
        width = _parse_fraction(cells[4])
        length = _parse_fraction(cells[5])
        hours = _safe_float(cells[6], 0.0)
        price = _safe_float(cells[7], 0.0)

        # Skip empty / zero-qty template rows (flipper blanks, etc.).
        if qty <= 0 and thickness <= 0 and width <= 0 and length <= 0:
            continue
        if qty <= 0:
            continue
        # Skip spare rows that have no real name and no default (blank extras).
        if not name or (not default_name and role == "steel_plate" and not raw_name):
            if role == "steel_plate" and not raw_name:
                continue

        role_key = _BMS_NAME_TO_ROLE.get(_norm(name), role)
        if role_key in _BMS_ROLE_LABELS and (not raw_name or role_key in {"tcp", "bcp", "id_holder", "od_holder", "id_pot", "od_pot"}):
            # Keep canonical BMS labels for the six pot-block slots.
            if default_name:
                name = default_name
        cu_in = 0.0
        if thickness > 0 and width > 0 and length > 0:
            cu_in = round(qty * thickness * width * length, 2)

        lines.append(
            {
                "component": name,
                "role": role_key,
                "qty": qty,
                "thickness": thickness,
                "width": width,
                "length": length,
                "hours": hours if hours > 0 else None,
                "cu_in": cu_in if cu_in > 0 else None,
                "price": round(price, 2) if price > 0 else 0.0,
                "price_source": f"quote_workbook:{sheet_name}:row{row_idx}",
                "source_sheet": sheet_name,
            }
        )
    return lines


def _load_steel_from_j000(wb_path: Path | None) -> list[dict]:
    """J000 Steel Order / Machining Sheet: A=Qty B=Name C=T E=W G=L."""
    if not wb_path:
        return []
    lines: list[dict] = []
    seen: set[str] = set()
    for sheet_name, row_idx, cells in _iter_workbook_rows(
        wb_path, ("Steel Order", "Machining Sheet", "Steel")
    ):
        if row_idx < 19:
            continue
        while len(cells) < 8:
            cells.append(None)
        name = _cell_str(cells[1])
        if not name:
            continue
        role = _BMS_NAME_TO_ROLE.get(_norm(name), "")
        # Also accept standard plate names via PLATE_SHEET_NAMES keywords.
        if not role:
            cn = _norm(name)
            for r, names in PLATE_SHEET_NAMES.items():
                if any(n in cn for n in names):
                    role = r
                    break
        if not role:
            # Still show named steel rows Module6121 wrote (extras like Flipper).
            role = "steel_plate"
        qty = _safe_float(cells[0], 1.0)
        if qty <= 0:
            continue
        thickness = _parse_fraction(cells[2])
        width = _parse_fraction(cells[4])
        length = _parse_fraction(cells[6])
        if thickness <= 0 or width <= 0 or length <= 0:
            continue
        key = _norm(name)
        if key in seen:
            continue
        seen.add(key)
        cu_in = round(qty * thickness * width * length, 2)
        lines.append(
            {
                "component": _canonical_steel_name(name, role),
                "role": role,
                "qty": qty,
                "thickness": thickness,
                "width": width,
                "length": length,
                "hours": None,
                "cu_in": cu_in,
                "price": 0.0,
                "price_source": f"steel_workbook:{sheet_name}:row{row_idx}",
                "source_sheet": sheet_name,
            }
        )
    return lines


def _load_steel_from_bom_match(job_dir: Path) -> list[dict]:
    """Fallback: XT_Export_BOM_Match_Report.csv QuoteName + CAD dims."""
    for name in (
        "XT_Export_BOM_Match_Report.csv",
        "documents/XT_Export_BOM_Match_Report.csv",
    ):
        rows = _read_csv_rows(job_dir / name)
        if not rows:
            continue
        lines: list[dict] = []
        for row in rows:
            qn = _cell_str(row.get("QuoteName") or row.get("quoteName"))
            if not qn:
                continue
            role = _BMS_NAME_TO_ROLE.get(_norm(qn), "")
            if not role:
                continue
            qty = _safe_float(row.get("Qty") or row.get("QTY") or 1, 1.0)
            thickness = _safe_float(row.get("CAD_Thickness") or row.get("BOM_Thickness"))
            width = _safe_float(row.get("CAD_Width") or row.get("BOM_Width"))
            length = _safe_float(row.get("CAD_Length") or row.get("BOM_Length"))
            if thickness <= 0 or width <= 0 or length <= 0:
                continue
            cu_in = round(qty * thickness * width * length, 2)
            lines.append(
                {
                    "component": qn,
                    "role": role,
                    "qty": qty,
                    "thickness": thickness,
                    "width": width,
                    "length": length,
                    "hours": None,
                    "cu_in": cu_in,
                    "price": 0.0,
                    "price_source": "bom_match_report",
                    "source_sheet": "",
                }
            )
        if lines:
            return lines
    return []


def load_pullcore_lines(job_dir: Path) -> list[dict]:
    """Pull cores & keys from Pullcore Prices.csv (Module6121 WritePullcorePriceFile)."""
    rows = load_job_pullcore(job_dir)
    lines: list[dict] = []
    for row in rows:
        name = _cell_str(
            row.get("Pull Core / Key")
            or row.get("Description")
            or row.get("Component")
        )
        if not name or name.upper() in ("TOTAL", "RATE ($/IN3)", "RATE"):
            continue
        if name.upper().startswith("RATE"):
            continue
        qty = _safe_float(row.get("Qty") or row.get("QTY") or 1, 1.0)
        thickness = _safe_float(row.get("Thickness"))
        width = _safe_float(row.get("Width"))
        length = _safe_float(row.get("Length"))
        cu_in = _safe_float(row.get("Cu In") or row.get("Cu. In.") or row.get("CuIn"))
        price = _safe_float(row.get("Price USD") or row.get("Price"))
        if cu_in <= 0 and thickness > 0 and width > 0 and length > 0:
            cu_in = round(qty * thickness * width * length, 2)
        lines.append(
            {
                "component": name,
                "role": "pullcore",
                "qty": qty,
                "thickness": thickness,
                "width": width,
                "length": length,
                "hours": None,
                "cu_in": cu_in if cu_in > 0 else None,
                "price": round(price, 2),
                "price_source": "job_csv:Pullcore Prices.csv",
                "material": _cell_str(row.get("Material")),
            }
        )
    return lines


def load_purchased_lines(job_dir: Path) -> list[dict]:
    """Purchased components from Purchased Components Quote.csv."""
    rows = load_job_purchased_quote(job_dir)
    lines: list[dict] = []
    for row in rows:
        desc = _cell_str(row.get("Description") or row.get("Component"))
        comp = _cell_str(row.get("Component") or desc)
        if not desc and not comp:
            continue
        if (desc or comp).upper() == "TOTAL":
            continue
        qty = _safe_float(row.get("QTY") or row.get("Qty") or 1, 1.0)
        unit = _safe_float(row.get("UnitPrice"))
        ext = _safe_float(row.get("Extended"))
        if ext <= 0 and unit > 0:
            ext = unit * max(qty, 1)
        lines.append(
            {
                "component": desc or comp,
                "role": "purchased_component",
                "qty": qty,
                "thickness": None,
                "width": None,
                "length": None,
                "hours": None,
                "cu_in": None,
                "price": round(ext, 2),
                "price_source": "job_csv:Purchased Components Quote.csv",
                "vendor": _cell_str(row.get("Vendor")),
                "part_number": _cell_str(row.get("PartNumber")),
                "unit_price": unit,
                "category": comp,
            }
        )
    return lines


def load_quote_summary(job_dir: Path) -> dict:
    """Best-effort summary totals from QuoteWorksheet (hours / price / commission)."""
    wb = _find_workbook(job_dir)
    summary = {
        "total_hours": None,
        "total_price_rough": None,
        "total_price_finish": None,
        "commission_pct": None,
        "commission_rough": None,
        "commission_finish": None,
        "grand_total_rough": None,
        "grand_total_finish": None,
    }
    if not wb:
        return summary

    for sheet_name, row_idx, cells in _iter_workbook_rows(
        wb, ("QuoteWorksheet", "Quote")
    ):
        if not cells:
            continue
        label = _norm(_cell_str(cells[0]))
        nums = [_safe_float(c) for c in cells[1:6] if _safe_float(c) != 0]
        if "total hours" in label and nums:
            summary["total_hours"] = nums[0]
        elif label == "commission" or label.startswith("commission"):
            # Often: Commission | 6% | $545 | $550
            pct = None
            for c in cells[1:4]:
                s = _cell_str(c)
                if s.endswith("%"):
                    pct = _safe_float(s.replace("%", ""))
            money = [_safe_float(c) for c in cells[1:5] if _safe_float(c) > 1]
            if pct is not None:
                summary["commission_pct"] = pct
            if len(money) >= 1:
                summary["commission_rough"] = money[0]
            if len(money) >= 2:
                summary["commission_finish"] = money[1]
        elif label == "total price":
            money = [_safe_float(c) for c in cells[1:5] if _safe_float(c) > 1]
            # First Total Price row is subtotal; later one (after commission) is grand.
            if summary["total_price_rough"] is None and money:
                summary["total_price_rough"] = money[0]
                if len(money) >= 2:
                    summary["total_price_finish"] = money[1]
            elif money:
                summary["grand_total_rough"] = money[0]
                if len(money) >= 2:
                    summary["grand_total_finish"] = money[1]
    return summary


def _match_shop_price(role: str, component_name: str, shop_rows: list[dict]) -> tuple[float, str]:
    """Return (unit_price, csv_component_matched)."""
    comp_norm = _norm(component_name)
    keywords = ROLE_CSV_KEYWORDS.get(role, [])
    best_price = 0.0
    best_match = ""
    is_plate_role = role.endswith("_plate") or role in ("rail", "pin_plate")

    for row in shop_rows:
        csv_comp = _norm(row.get("Component") or "")
        price = _safe_float(row.get("UnitPrice"))
        if price <= 0 or not csv_comp:
            continue

        # Plates must not pick up pin/bushing hardware rows.
        if is_plate_role and any(x in csv_comp for x in ("pin", "bushing", "strap", "ring", "insulation")):
            if "plate" not in csv_comp:
                continue
        if role == "leader_pin" and "pin" not in csv_comp:
            continue
        if role == "leader_pin_bushing" and "bush" not in csv_comp:
            continue

        if csv_comp in comp_norm or comp_norm in csv_comp:
            return price, row.get("Component", "")

        for kw in keywords:
            if len(kw) < 3:
                continue
            if kw in csv_comp or (kw in comp_norm and not is_plate_role):
                if price > best_price:
                    best_price = price
                    best_match = row.get("Component", "")

    return best_price, best_match


def _match_job_purchased(
    component_name: str, job_rows: list[dict]
) -> tuple[float, str]:
    comp_norm = _norm(component_name)
    for row in job_rows:
        csv_comp = _norm(row.get("Component") or "")
        if not csv_comp:
            continue
        unit = _safe_float(row.get("UnitPrice"))
        qty = _safe_float(row.get("QTY") or row.get("Qty") or 1, 1.0)
        ext = _safe_float(row.get("Extended"))
        if ext > 0:
            return ext, row.get("Component", "")
        if unit > 0:
            return unit * max(qty, 1), row.get("Component", "")
        if csv_comp in comp_norm or comp_norm in csv_comp:
            return unit * max(qty, 1), row.get("Component", "")
    return 0.0, ""


def _parse_fraction(val) -> float:
    s = str(val).strip()
    if not s:
        return 0.0
    m = re.match(r"^(\d+)\s+(\d+)/(\d+)$", s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / int(m.group(3))
    m = re.match(r"^(\d+)/(\d+)$", s)
    if m:
        return int(m.group(1)) / int(m.group(2))
    return _safe_float(s)


def read_sheet_dimensions(job_dir: Path) -> dict[str, dict]:
    """Parse quote/steel Excel for plate names and T/W/L (stock sizes on quote sheet)."""
    wb_path = _find_workbook(job_dir)
    if not wb_path:
        return {}
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return {}

    out: dict[str, dict] = {}
    try:
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    except Exception:
        return {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=250, values_only=True):
            if not row:
                continue
            cells = [str(c).strip() if c is not None else "" for c in row]
            for role, names in PLATE_SHEET_NAMES.items():
                if role in out:
                    continue
                for cell in cells:
                    cn = _norm(cell)
                    if any(n in cn for n in names):
                        nums = [_parse_fraction(c) for c in cells if _parse_fraction(c) > 0]
                        if len(nums) >= 3:
                            out[role] = {
                                "thickness": nums[0],
                                "width": nums[1],
                                "length": nums[2],
                                "sheet_name": cell,
                                "source_sheet": sheet_name,
                            }
                        break
    wb.close()
    return out


def price_for_part(
    row: dict,
    shop_rows: list[dict],
    job_purchased: list[dict],
    sheet_dims: dict[str, dict],
) -> dict:
    """Return pricing dict with price, source, and sheet dimensions when available."""
    role = row.get("role", "")
    component = row.get("Component") or row.get("component") or ""
    quote_flag = bool(row.get("quote") or row.get("Quote"))

    dims = sheet_dims.get(role, {})
    thickness = dims.get("thickness") or _safe_float(row.get("Thickness"))
    width = dims.get("width") or _safe_float(row.get("Width"))
    length = dims.get("length") or _safe_float(row.get("Length"))

    if not quote_flag:
        return {
            "price": 0.0,
            "price_source": "",
            "thickness": thickness,
            "width": width,
            "length": length,
        }

    price = 0.0
    source = ""

    if job_purchased:
        price, match = _match_job_purchased(component, job_purchased)
        if price > 0:
            source = f"job_csv:{match}"

    if price <= 0 and shop_rows:
        price, match = _match_shop_price(role, component, shop_rows)
        if price > 0:
            source = f"shop_csv:{match}"

    if price <= 0:
        source = "no_csv_price"

    return {
        "price": round(price, 2),
        "price_source": source,
        "thickness": thickness,
        "width": width,
        "length": length,
    }
